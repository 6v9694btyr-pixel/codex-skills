#!/usr/bin/env python3
"""Fill a TikTok Seller Center listing template from a structured payload."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required to edit TikTok Seller Center .xlsx templates. "
        "Install openpyxl or run this script with the Codex Desktop bundled Python runtime."
    ) from exc


ALLOWED_CATEGORIES = {
    "Costume Jewelry & Accessories/Bracelets & Bangles",
    "Costume Jewelry & Accessories/Keychains",
    "Eyewear/Sunglasses",
}

IMAGE_SLOTS = [
    "main_image",
    "image_2",
    "image_3",
    "image_4",
    "image_5",
    "image_6",
    "image_7",
    "image_8",
    "image_9",
]

IMAGE_FILENAMES = {
    "main_image": "main_image.png",
    "image_2": "image_2.png",
    "image_3": "image_3.png",
    "image_4": "image_4.png",
    "image_5": "image_5.png",
    "image_6": "image_6.png",
    "image_7": "image_7.png",
    "image_8": "image_8.png",
    "image_9": "image_9.png",
}

MATERIALS_BY_CATEGORY = {
    "Costume Jewelry & Accessories/Keychains": {
        "Fabric",
        "Plastic",
        "Metal Coating",
    },
    "Costume Jewelry & Accessories/Bracelets & Bangles": {
        "Fabric",
        "Plastic",
        "Metal Coating",
        "Wood",
        "Leather",
    },
    "Eyewear/Sunglasses": {
        "Plastic",
        "Metal Coating",
    },
}

DEFAULT_COMMERCE_FIELDS = {
    "parcel_weight": "100",
    "parcel_length": "16",
    "parcel_width": "13",
    "parcel_height": "5",
    "price": "178",
    "quantity": "100",
}

FIELD_ALIASES = {
    "category": ["category", "productcategory", "categoryid", "categorypath"],
    "brand": ["brand", "brandname"],
    "product_name": ["productname", "producttitle", "title", "itemname"],
    "product_description": ["productdescription", "description", "itemdescription"],
    "main_image": ["mainimage", "primaryimage", "productmainimage"],
    "image_2": ["image2", "productimage2", "additionalimage2"],
    "image_3": ["image3", "productimage3", "additionalimage3"],
    "image_4": ["image4", "productimage4", "additionalimage4"],
    "image_5": ["image5", "productimage5", "additionalimage5"],
    "image_6": ["image6", "productimage6", "additionalimage6"],
    "image_7": ["image7", "productimage7", "additionalimage7"],
    "image_8": ["image8", "productimage8", "additionalimage8"],
    "image_9": ["image9", "productimage9", "additionalimage9"],
    "gtin_type": ["gtintype", "productidentifiercodetype", "identifiercodetype"],
    "gtin_code": ["gtincode", "productidentifiercode", "identifiercode"],
    "property_name_1": ["propertyname1", "variationname1", "salesattributename1"],
    "property_value_1": ["propertyvalue1", "variationvalue1", "salesattributevalue1"],
    "property_1_image": ["property1image", "variationimage1", "salesattributeimage1"],
    "property_name_2": ["propertyname2", "variationname2", "salesattributename2"],
    "property_value_2": ["propertyvalue2", "variationvalue2", "salesattributevalue2"],
    "parcel_weight": ["parcelweight", "packageweight", "packageweightg"],
    "parcel_length": ["parcellength", "packagelength", "packagelengthcm"],
    "parcel_width": ["parcelwidth", "packagewidth", "packagewidthcm"],
    "parcel_height": ["parcelheight", "packageheight", "packageheightcm"],
    "delivery": ["delivery", "deliveryoptions", "shippingoption"],
    "price": ["price", "retailprice", "retailpricelocalcurrency"],
    "quantity": ["quantity", "stock", "inventory"],
    "seller_sku": ["sellersku", "sku"],
}

SPANISH_HINTS = {
    "el",
    "la",
    "los",
    "las",
    "para",
    "con",
    "producto",
    "regalo",
    "llaves",
    "mochila",
    "bolso",
    "uso",
    "diario",
    "color",
    "puede",
}

RISK_PATTERN = re.compile(
    r"\b(100\s*%|garantizad[oa]|garantia total|oficial|autorizad[oa]|certificad[oa]|"
    r"original de marca|cura|curar|tratamiento|terapia|medic[oa]|salud|alivia|"
    r"permanente|para siempre|irrompible|indestructible|nunca se rompe|el mejor|numero 1|número 1)\b",
    re.IGNORECASE,
)


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[\*\:\-\/_]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def find_template_headers(ws: Any) -> tuple[int, dict[str, int]]:
    alias_lookup: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_lookup[norm(alias)] = field

    best_row = 0
    best_mapping: dict[str, int] = {}
    max_scan_row = min(ws.max_row, 20)
    for row_idx in range(1, max_scan_row + 1):
        mapping: dict[str, int] = {}
        for cell in ws[row_idx]:
            cell_norm = norm(cell.value)
            if not cell_norm:
                continue
            field = alias_lookup.get(cell_norm)
            if field is None:
                for alias_norm, alias_field in alias_lookup.items():
                    if len(alias_norm) >= 6 and alias_norm in cell_norm:
                        field = alias_field
                        break
            if field and field not in mapping:
                mapping[field] = cell.column
        if len(mapping) > len(best_mapping):
            best_row = row_idx
            best_mapping = mapping

    if len(best_mapping) < 4:
        raise ValueError(
            "Could not detect enough template headers in the Template worksheet. "
            "Expected fields such as category, brand, product_name, and main_image."
        )
    return best_row, best_mapping


def require(condition: bool, message: str, errors: list[str]) -> None:
    if not condition:
        errors.append(message)


def looks_spanish(text: str) -> bool:
    tokens = set(re.findall(r"[a-záéíóúñü]+", text.lower()))
    return len(tokens & SPANISH_HINTS) >= 4


def image_value(slot: str, payload: dict[str, Any]) -> tuple[str, bool]:
    urls = payload.get("image_urls") or {}
    url = str(urls.get(slot) or "").strip()
    if url:
        return url, False
    return f"PLEASE_UPLOAD_{IMAGE_FILENAMES[slot]}", True


def clean_property_value(category: str, property_name: str, value: str) -> str:
    if property_name != "Material":
        return value.strip()
    allowed = MATERIALS_BY_CATEGORY.get(category, set())
    value = value.strip()
    return value if value in allowed else "Plastic"


def build_fields(payload: dict[str, Any], variant: dict[str, Any] | None = None) -> dict[str, str]:
    variant = variant or {}
    category = str(payload.get("category") or "").strip()
    analysis = payload.get("analysis") or {}
    default_property_name = "Color" if category == "Eyewear/Sunglasses" else "Material"
    property_name_1 = str(variant.get("property_name_1") or payload.get("property_name_1") or default_property_name).strip()
    material_guess = str(
        variant.get("property_value_1")
        or payload.get("property_value_1")
        or analysis.get("material")
        or "Plastic"
    ).strip()
    property_value_1 = clean_property_value(category, property_name_1, material_guess)

    fields = {
        "category": category,
        "brand": str(payload.get("brand") or "No brand").strip(),
        "product_name": str(payload.get("product_name") or "").strip(),
        "product_description": str(payload.get("product_description") or "").strip(),
        "gtin_type": str(payload.get("gtin_type") or "").strip(),
        "gtin_code": str(payload.get("gtin_code") or "").strip(),
        "property_name_1": property_name_1,
        "property_value_1": property_value_1,
        "property_1_image": str(variant.get("property_1_image") or payload.get("property_1_image") or "").strip(),
        "property_name_2": str(variant.get("property_name_2") or payload.get("property_name_2") or "").strip(),
        "property_value_2": str(variant.get("property_value_2") or payload.get("property_value_2") or "").strip(),
        "parcel_weight": str(
            variant.get("parcel_weight") or payload.get("parcel_weight") or DEFAULT_COMMERCE_FIELDS["parcel_weight"]
        ).strip(),
        "parcel_length": str(
            variant.get("parcel_length") or payload.get("parcel_length") or DEFAULT_COMMERCE_FIELDS["parcel_length"]
        ).strip(),
        "parcel_width": str(
            variant.get("parcel_width") or payload.get("parcel_width") or DEFAULT_COMMERCE_FIELDS["parcel_width"]
        ).strip(),
        "parcel_height": str(
            variant.get("parcel_height") or payload.get("parcel_height") or DEFAULT_COMMERCE_FIELDS["parcel_height"]
        ).strip(),
        "delivery": str(variant.get("delivery") or payload.get("delivery") or "").strip(),
        "price": str(variant.get("price") or payload.get("price") or DEFAULT_COMMERCE_FIELDS["price"]).strip(),
        "quantity": str(variant.get("quantity") or payload.get("quantity") or DEFAULT_COMMERCE_FIELDS["quantity"]).strip(),
        "seller_sku": str(variant.get("seller_sku") or payload.get("seller_sku") or "").strip(),
    }

    placeholder_slots: list[str] = []
    for slot in IMAGE_SLOTS:
        value, is_placeholder = image_value(slot, payload)
        fields[slot] = value
        if is_placeholder:
            placeholder_slots.append(slot)

    if not fields["property_1_image"] and not fields["main_image"].startswith("PLEASE_UPLOAD_"):
        fields["property_1_image"] = fields["main_image"]

    payload["_placeholder_slots"] = placeholder_slots
    return fields


def validate_payload(payload: dict[str, Any], fields: dict[str, str], output_xlsx: Path) -> list[str]:
    errors: list[str] = []
    require(bool(fields["product_name"]), "product_name is empty.", errors)
    require(len(fields["product_name"]) <= 300, "product_name exceeds 300 characters.", errors)
    require(bool(fields["product_description"]), "product_description is empty.", errors)
    require(fields["category"] in ALLOWED_CATEGORIES, "category is not one of the allowed TikTok categories.", errors)
    require(fields["brand"] == "No brand", "brand must be exactly 'No brand'.", errors)
    require(bool(fields["main_image"]), "main_image is empty.", errors)
    filled_images = [slot for slot in IMAGE_SLOTS if fields.get(slot)]
    require(len(filled_images) >= 5, "At least 5 image fields must be filled.", errors)
    require(looks_spanish(fields["product_description"]), "product_description does not look like Spanish.", errors)
    risky_text = f"{fields['product_name']}\n{fields['product_description']}"
    require(RISK_PATTERN.search(risky_text) is None, "Risky or exaggerated claim detected.", errors)
    require(output_xlsx.suffix.lower() == ".xlsx", "Output workbook must be .xlsx.", errors)
    return errors


def copy_images(payload: dict[str, Any], payload_path: Path, output_dir: Path) -> None:
    images_dir = output_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)
    image_files = payload.get("image_files") or {}
    for slot, filename in IMAGE_FILENAMES.items():
        raw_path = str(image_files.get(slot) or "").strip()
        if not raw_path:
            continue
        src = Path(raw_path)
        if not src.is_absolute():
            src = payload_path.parent / src
        if src.exists() and src.is_file():
            dst = images_dir / filename
            if src.resolve() != dst.resolve():
                shutil.copy2(src, dst)


def write_manifest(payload: dict[str, Any], output_dir: Path) -> None:
    placeholder_slots = payload.get("_placeholder_slots") or []
    if not placeholder_slots:
        return
    manifest_path = output_dir / "image_upload_manifest.csv"
    purposes = {
        "main_image": "White-background main product image for main_image",
        "image_2": "Lifestyle use scene",
        "image_3": "Detail close-up",
        "image_4": "Size-feel reference image",
        "image_5": "Selling point image 1",
        "image_6": "Multi-scene use image",
        "image_7": "Gift scene image",
        "image_8": "Simple package or set layout",
        "image_9": "TikTok-style conversion mood image",
    }
    with manifest_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["image_slot", "local_file_name", "recommended_upload_name", "purpose"],
        )
        writer.writeheader()
        for slot in IMAGE_SLOTS:
            if slot not in placeholder_slots:
                continue
            filename = IMAGE_FILENAMES[slot]
            writer.writerow(
                {
                    "image_slot": slot,
                    "local_file_name": f"images/{filename}",
                    "recommended_upload_name": filename,
                    "purpose": purposes[slot],
                }
            )


def md_escape(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def write_summary(payload: dict[str, Any], fields: dict[str, str], output_dir: Path) -> None:
    analysis = payload.get("analysis") or {}
    image_plan = payload.get("image_plan") or []
    image_prompts = payload.get("image_prompts") or []

    lines: list[str] = []
    lines.append("# TikTok Shop Mexico Listing Summary")
    lines.append("")
    lines.append("## 产品分析")
    if analysis:
        for key, value in analysis.items():
            lines.append(f"- {key}: {md_escape(value)}")
    else:
        lines.append("- No analysis details supplied.")
    lines.append("")
    lines.append("## 选择的类目")
    lines.append(fields["category"])
    lines.append("")
    lines.append("## 商品标题")
    lines.append(fields["product_name"])
    lines.append("")
    lines.append("## 商品描述")
    lines.append(fields["product_description"])
    lines.append("")
    lines.append("## 9 张主图说明")
    if image_plan:
        for item in image_plan:
            lines.append(
                f"- {item.get('slot', '')}: {item.get('cn') or item.get('purpose') or ''}"
            )
    else:
        for slot in IMAGE_SLOTS:
            lines.append(f"- {slot}: {IMAGE_FILENAMES[slot]}")
    lines.append("")
    lines.append("## 9 条图片提示词")
    if image_prompts:
        for item in image_prompts:
            lines.append(f"### {item.get('slot', '')}")
            if item.get("cn"):
                lines.append(md_escape(item.get("cn")))
            if item.get("prompt_en"):
                lines.append("")
                lines.append("```text")
                lines.append(md_escape(item.get("prompt_en")))
                lines.append("```")
    else:
        lines.append("No image prompts supplied.")
    lines.append("")
    lines.append("## Excel 填写字段摘要")
    for key in [
        "category",
        "brand",
        "product_name",
        "main_image",
        "image_2",
        "image_3",
        "image_4",
        "image_5",
        "image_6",
        "image_7",
        "image_8",
        "image_9",
        "gtin_type",
        "gtin_code",
        "property_name_1",
        "property_value_1",
        "property_1_image",
        "property_name_2",
        "property_value_2",
        "parcel_weight",
        "parcel_length",
        "parcel_width",
        "parcel_height",
        "delivery",
        "price",
        "quantity",
        "seller_sku",
    ]:
        lines.append(f"- {key}: {fields.get(key, '')}")
    lines.append("")
    (output_dir / "listing_summary.md").write_text("\n".join(lines), encoding="utf-8")


def find_data_start_row(ws: Any, header_row: int, header_map: dict[str, int]) -> int:
    fallback = max(header_row + 1, 6)
    category_col = header_map.get("category")
    if not category_col or not ws.data_validations:
        return fallback

    candidates: list[int] = []
    for dv in ws.data_validations.dataValidation:
        for cell_range in dv.cells.ranges:
            min_col, min_row, max_col, _max_row = range_boundaries(str(cell_range))
            if min_col <= category_col <= max_col and min_row > header_row:
                candidates.append(min_row)
    return min(candidates) if candidates else fallback


def fill_workbook(template: Path, rows: list[dict[str, str]], output_xlsx: Path) -> None:
    workbook = load_workbook(template)
    if "Template" not in workbook.sheetnames:
        raise ValueError("Workbook does not contain a 'Template' worksheet.")
    ws = workbook["Template"]
    header_row, header_map = find_template_headers(ws)
    data_row = find_data_start_row(ws, header_row, header_map)

    for offset, fields in enumerate(rows):
        row_idx = data_row + offset
        for field, value in fields.items():
            column = header_map.get(field)
            if column is not None:
                ws.cell(row=row_idx, column=column).value = value

    workbook.save(output_xlsx)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--template", required=True, help="Path to TikTok Seller Center .xlsx template.")
    parser.add_argument("--payload", required=True, help="Path to listing_payload.json.")
    parser.add_argument("--output-dir", required=True, help="Directory for completed workbook and summary files.")
    parser.add_argument("--output-xlsx", default="completed_tiktok_listing.xlsx", help="Output workbook filename.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    template = Path(args.template).expanduser().resolve()
    payload_path = Path(args.payload).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = output_dir / args.output_xlsx

    if not template.exists():
        raise FileNotFoundError(f"Template not found: {template}")
    if template.suffix.lower() != ".xlsx":
        raise ValueError("Template must be an .xlsx file.")
    if not payload_path.exists():
        raise FileNotFoundError(f"Payload not found: {payload_path}")

    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    variants = payload.get("variants") or []
    row_fields = [build_fields(payload, variant) for variant in variants] if variants else [build_fields(payload)]
    fields = row_fields[0]
    errors = validate_payload(payload, fields, output_xlsx)
    if errors:
        for error in errors:
            print(f"ERROR: {error}", file=sys.stderr)
        return 2

    copy_images(payload, payload_path, output_dir)
    fill_workbook(template, row_fields, output_xlsx)
    write_manifest(payload, output_dir)
    write_summary(payload, fields, output_dir)
    shutil.copy2(payload_path, output_dir / "listing_payload.json")

    print(f"Wrote {output_xlsx}")
    print(f"Wrote {output_dir / 'listing_summary.md'}")
    if (output_dir / "image_upload_manifest.csv").exists():
        print(f"Wrote {output_dir / 'image_upload_manifest.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
